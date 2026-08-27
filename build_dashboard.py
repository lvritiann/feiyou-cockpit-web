# -*- coding: utf-8 -*-
"""
build_dashboard.py — 把「非油销售汇总驾驶舱V1.xlsx」图表数据源里的数据
重做成纯静态网页 (web/index.html) + 数据层 (web/data.json)。

设计要点:
- 全部用【标签扫描】定位「图表数据源」的 块1~块6 与「按月汇总各站」的分类明细块，
  绝不硬编码行号(此前硬编码曾导致尾部站点错位，见项目记忆)。
- 提取覆盖: 15 站 × 12 月 金额/毛利(含+不含) / 各站年任务与累计完成 / 公司月度
  金额+环比 / 品类构成(5类) / 双口径 / 【品类明细】(5品类+非油合计 × 15站 × 12月 金额+毛利)。
- 生成的 index.html: 数据内嵌 + 联动(站点选择器 + 点击图表联动筛选) + 站点明细弹窗
  (品类×12月金额/毛利堆叠柱 + 构成饼 + 合计表) + porcelain 青瓷蓝主题，加载飞快。
- 公司级图表(构成/双口径/任务/月度)在选中某站后自动切换为该站数据。
- 每日自动更新由 GitHub Action(10点) 重跑本脚本(读 data.json) + AirScript 云端
  推送 data.json 共同完成。

用法:
    python build_dashboard.py            # 读 Excel -> data.json + index.html
    python build_dashboard.py --from-json # 只读 data.json -> 重渲染 index.html (Action 用)
依赖: openpyxl
"""
import os, json, sys, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
DASH = os.path.join(HERE, "..", "非油销售汇总驾驶舱V1.xlsx")
DATA_JSON = os.path.join(HERE, "data.json")
INDEX = os.path.join(HERE, "index.html")

STATION_ORDER = ["秦岭", "宁陕", "洋县", "汉中", "富平", "韩城", "富县", "南沙",
                 "华山", "白河", "旬阳", "略阳", "玉华宫", "照金", "天汉水城"]
CAT_DETAIL = ["汽车用品", "便利百货", "香烟零售", "烟草批发", "咖啡"]
CAT_DETAIL_ALL = CAT_DETAIL + ["非油合计"]
TASK_TOTAL = 1130  # 全年任务(万元)


def load_excel():
    import openpyxl
    return openpyxl.load_workbook(DASH, data_only=False)


def find_row(ws, kw, start=1, end=None):
    end = end or ws.max_row
    for r in range(start, end + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and kw in v:
            return r
    return None


def station_rows(ws, after_title):
    """标题行之后, 站名所在的行号映射(扫描到非站名/合计/下一块为止)。"""
    rows = {}
    for r in range(after_title + 2, after_title + 40):
        v = ws.cell(r, 1).value
        if v in STATION_ORDER:
            rows[v] = r
        elif isinstance(v, str) and ("合计" in v or "块" in v):
            break
    return rows


def total_row(ws, after_title):
    for r in range(after_title + 2, after_title + 40):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "合计" in v:
            return r
    return None


def build_data():
    wb = load_excel()
    cd = wb["图表数据源"]
    ms = wb["按月汇总各站"]
    dd = wb["按日汇总"]  # 近30天日变化数据源

    # 探测最新有数据月(块5 含烟草合计行, 1月=col2...12月=col13): 供所有月度数据截断
    b5_probe = find_row(cd, "块5 ·")
    r_tob_probe = find_row(cd, "含烟草合计", start=b5_probe)
    tob12 = [cd.cell(r_tob_probe, 1 + m).value for m in range(1, 13)]
    LAST_MONTH = 12
    for m in range(11, -1, -1):
        if tob12[m] not in (None, ""):
            LAST_MONTH = m + 1
            break

    b1 = find_row(cd, "块1 ·"); r1 = station_rows(cd, b1)
    b2 = find_row(cd, "块2 ·"); r2 = station_rows(cd, b2)
    b3 = find_row(cd, "块3 ·"); r3 = station_rows(cd, b3)
    b4 = find_row(cd, "块4 ·"); r4 = station_rows(cd, b4)

    def monthly(rmap):
        return {st: [cd.cell(r, 1 + m).value for m in range(1, LAST_MONTH + 1)] for st, r in rmap.items()}

    monthly_data = {
        "tob": monthly(r1), "non": monthly(r2),
        "tobProfit": monthly(r3), "nonProfit": monthly(r4),
    }

    # ── 近30天日变化: 读「按日汇总」每站每日 金额(col2-16)+毛利(col18-32) + 合计(col17/33) ──
    daily30 = build_daily30(dd)

    # 块6: 各站年任务与累计完成(万元)
    b6 = find_row(cd, "块6 ·")
    cr = station_rows(cd, b6)
    stations = []
    for st in STATION_ORDER:
        r = cr[st]
        stations.append({
            "name": st,
            "yearTask": cd.cell(r, 2).value,
            "compTob": cd.cell(r, 3).value,
            "rateTob": cd.cell(r, 4).value,
            "compNon": cd.cell(r, 5).value,
            "rateNon": cd.cell(r, 6).value,
        })

    # 块5: 公司月度 + 环比（保留，供公司视角月度图）
    b5 = find_row(cd, "块5 ·")
    r_tob = find_row(cd, "含烟草合计", start=b5)
    r_non = find_row(cd, "不含烟草合计", start=b5)
    r_tobm = find_row(cd, "含烟草环比", start=b5)
    r_nonm = find_row(cd, "不含烟草环比", start=b5)
    # 月份轴截断到最新有数据月(LAST_MONTH)
    n_m = LAST_MONTH
    company = {
        "monthLabels": [f"{m}月" for m in range(1, n_m + 1)],
        "tob": [cd.cell(r_tob, 1 + m).value for m in range(1, n_m + 1)],
        "non": [cd.cell(r_non, 1 + m).value for m in range(1, n_m + 1)],
        "tobMom": [cd.cell(r_tobm, 1 + m).value for m in range(1, n_m + 1)],
        "nonMom": [cd.cell(r_nonm, 1 + m).value for m in range(1, n_m + 1)],
    }

    # ── 分类明细块：按月汇总各站「各站逐月分类明细」(5品类 + 非油合计) ×15站×12月 金额+毛利 ──
    def cat_titles():
        titles = {}
        for r in range(40, ms.max_row + 1):
            v = ms.cell(r, 1).value
            if isinstance(v, str) and "·" in v:
                name = v.split("·")[0].strip()
                if name in CAT_DETAIL_ALL:
                    titles[name] = r
        return titles

    def cat_matrix(title_row):
        """标题行后按序收集站名行：前 15 行为金额区、后 15 行为毛利区。"""
        st_rows = []
        for r in range(title_row + 1, title_row + 45):
            v = ms.cell(r, 1).value
            if v in STATION_ORDER:
                st_rows.append((v, r))
        amt = {name: [ms.cell(r, 1 + m).value for m in range(1, LAST_MONTH + 1)] for name, r in st_rows[:15]}
        prof = {name: [ms.cell(r, 1 + m).value for m in range(1, LAST_MONTH + 1)] for name, r in st_rows[15:30]}
        return amt, prof

    ct = cat_titles()
    cat_amount, cat_profit = {}, {}
    for c in CAT_DETAIL_ALL:
        amt, prof = cat_matrix(ct[c])
        cat_amount[c] = amt
        cat_profit[c] = prof
    category_detail = {"cats": CAT_DETAIL_ALL, "amount": cat_amount, "profit": cat_profit}

    # 公司品类年度合计（5 品类，万元）—— 由分类明细块汇总，保证与明细一致
    comp5 = []
    for c in CAT_DETAIL:
        total = sum(v or 0 for st in STATION_ORDER for v in cat_amount[c].get(st, [0] * 12))
        comp5.append(round(total / 1e4, 2))
    composition = {"cats": CAT_DETAIL, "tob": comp5}

    # 双口径 + 公司总额: 各块合计行 N 列
    n_tob = total_row(cd, b1); n_non = total_row(cd, b2)
    n_tobp = total_row(cd, b3); n_nonp = total_row(cd, b4)
    totals = {
        "tobAmount": round((cd.cell(n_tob, 14).value or 0) / 1e4, 2),
        "nonAmount": round((cd.cell(n_non, 14).value or 0) / 1e4, 2),
        "tobProfit": round((cd.cell(n_tobp, 14).value or 0) / 1e4, 2),
        "nonProfit": round((cd.cell(n_nonp, 14).value or 0) / 1e4, 2),
    }
    dual = {
        "cats": ["金额", "毛利"],
        "tob": [totals["tobAmount"], totals["tobProfit"]],
        "non": [totals["nonAmount"], totals["nonProfit"]],
    }

    # ── 预测模型(目标达成短信) → forecast 节点 ──
    forecast = build_forecast(wb)

    return {
        "generatedAt": datetime.date.today().isoformat(),
        "taskTotal": TASK_TOTAL,
        "stations": stations,
        "monthly": monthly_data,
        "company": company,
        "composition": composition,
        "categoryDetail": category_detail,
        "dual": dual,
        "totals": totals,
        "forecast": forecast,
        "daily30": daily30,
    }


def build_daily30(ws):
    """读「按日汇总」近30天(截至最新有数据日) 每站每日 金额/毛利 + 公司合计。
    ws: 「按日汇总」sheet。列结构(1-based): 1=日期, 2-16=15站金额, 17=金额合计,
       18-32=15站毛利, 33=毛利合计, 45=辅助日期。
    """
    import datetime as _dt
    # 找最新有数据的日期行(金额合计>0)
    rows = []
    for r in range(4, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, (_dt.datetime, _dt.date)):
            continue
        amt_total = ws.cell(r, 17).value
        if isinstance(amt_total, (int, float)) and amt_total > 0:
            rows.append(r)
    if not rows:
        return {"dates": [], "amount": {}, "profit": {}}
    end_r = rows[-1]
    start_r = rows[-30] if len(rows) >= 30 else rows[0]

    dates, amount, profit = [], {}, {}
    for i, st in enumerate(STATION_ORDER):
        amount[st] = []
        profit[st] = []
    amount["合计"] = []
    profit["合计"] = []
    for r in range(start_r, end_r + 1):
        d = ws.cell(r, 1).value
        if isinstance(d, _dt.datetime):
            dates.append(d.strftime("%m-%d"))
        elif isinstance(d, _dt.date):
            dates.append(d.strftime("%m-%d"))
        else:
            dates.append("")
        for i, st in enumerate(STATION_ORDER):
            amount[st].append(ws.cell(r, 2 + i).value)
            profit[st].append(ws.cell(r, 18 + i).value)
        amount["合计"].append(ws.cell(r, 17).value)
        profit["合计"].append(ws.cell(r, 33).value)
    return {"dates": dates, "amount": amount, "profit": profit}


def build_forecast(wb):
    """读取「预测模型」Sheet 参数 + 「目标达成」Sheet 预测短信，供『预测』标签页使用。"""
    pm = wb["预测模型"]
    L = {}
    for r in range(1, pm.max_row + 1):
        a = pm.cell(r, 1).value
        if isinstance(a, str) and a.strip():
            L[a.strip()] = pm.cell(r, 2).value

    def numv(k):
        v = L.get(k)
        try:
            return None if v is None else round(float(v), 4)
        except (TypeError, ValueError):
            return None

    # 短信正文:「目标达成」找到标题行后, 正文在下一行 A 列(或本行 B 列)
    sms = ""
    da = wb["目标达成"]
    for r in range(1, da.max_row + 1):
        a = da.cell(r, 1).value
        if isinstance(a, str) and "【每日销量预测短信】" in a:
            for cand in (da.cell(r + 1, 1).value, da.cell(r, 2).value):
                if isinstance(cand, str) and len(cand) > 50:
                    sms = cand.strip()
                    break
            break

    as_of = L.get("截至日期")
    if hasattr(as_of, "strftime"):
        as_of = as_of.strftime("%Y-%m-%d")
    pred_end = L.get("预测截止")
    if hasattr(pred_end, "strftime"):
        pred_end = pred_end.strftime("%Y-%m-%d")
    huashan = L.get("华山停业起")
    if hasattr(huashan, "strftime"):
        huashan = huashan.strftime("%Y-%m-%d")

    # 超额% 存的是小数(0.1956), 转成百分比数(19.56)
    def pct(v):
        return None if v is None else round(float(v) * 100, 2)

    # 模型里 毛利率/红线 已是小数形式(0.2408=24.08%), 直接乘100显示
    def pct100(v):
        return None if v is None else round(float(v) * 100, 2)

    return {
        "asOf": as_of,
        "predEnd": pred_end,
        "huashanFrom": huashan,
        "remainingMonths": numv("剩余月数"),
        "task": numv("全年任务万") or TASK_TOTAL,
        "kpis": {
            "累计含去化销售额万": numv("累计含去化销售额万"),
            "自然口径累计万": numv("自然口径累计万"),
            "整体毛利率": pct100(numv("整体毛利率")),
            "华山扣除万": numv("华山扣除万"),
        },
        "scenarios": {
            "labels": ["全年烟草去化", "剩余可去化", "全口径全年", "公司整体毛利", "期末香烟库存"],
            "A": [numv("X_A最大去化万(23.5%)"), numv("剩余可去化A万"), numv("全口径A万"),
                  numv("公司毛利A万"), numv("期末库存A万(23.5%)")],
            "B": [numv("X_B最大去化万(24%)"), numv("剩余可去化B万"), numv("全口径B万"),
                  numv("公司毛利B万"), numv("期末库存B万(24%)")],
        },
        "overrun": {"A": pct(numv("超额A%")), "B": pct(numv("超额B%"))},
        "extra": {
            "当前香烟库存万": numv("当前库存万"),
            "库存约束去化上限万": numv("库存约束去化上限X_cap万"),
            "库存是否封顶": numv("库存是否封顶(1=是,0=否)"),
            "剩余购进万": numv("剩余购进万"),
            "自然毛利率": pct100(numv("自然毛利率")),
            "红线A": pct100(numv("公司整体毛利红线A")),
            "红线B": pct100(numv("红线B")),
        },
        "sms": sms,
    }


HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>陕西高速延长石油有限责任公司非油销售汇总驾驶舱 · 图表视图</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
<script>
if (typeof echarts === 'undefined') {
  document.write('<scr'+'ipt src="https://lib.baomitu.com/echarts/5.5.0/echarts.min.js"><\/scr'+'ipt>');
}
</script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
  :root{
    --bg:#0d1320; --card:#162032; --ink:#eef2ff; --txt:#d7e0f0; --mut:#8fa1c5;
    --faint:#64748b; --grid:#2b3752; --line:#232f4a;
    --data:#3b82f6; --data2:#60a5fa; --hero:#3b82f6; --accent:#fbbf24;
    --good:#34d399;
  }
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);color:var(--txt);font-family:'Inter','Microsoft YaHei','PingFang SC',sans-serif;
    padding:26px 22px 40px;-webkit-font-smoothing:antialiased}
  .wrap{max-width:none;margin:0 auto;width:100%}
  header{margin-bottom:18px}
  header h1{font-size:22px;font-weight:700;color:var(--ink);letter-spacing:-.02em}
  header p{color:var(--mut);font-size:12.5px;margin-top:5px}
  header p.byline{color:var(--faint);font-size:11px;margin-top:3px}
  .upd{color:var(--faint);font-size:11px;margin-top:2px}

  /* 联动站点选择器 */
  .stations{display:flex;flex-wrap:wrap;gap:7px;margin:14px 0 18px}
  .chip{padding:6px 13px;border-radius:999px;border:1px solid var(--line);background:var(--card);
    color:var(--mut);font-size:12.5px;cursor:pointer;transition:.18s;user-select:none}
  .chip:hover{border-color:var(--data2);color:var(--data)}
  .chip.active{background:var(--hero);color:#fff;border-color:var(--hero);font-weight:600}
  .chip.all.active{background:var(--accent);border-color:var(--accent);color:#1e293b}

  /* KPI 行 */
  .kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:16px}
  .kpi{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:16px 18px}
  .kpi .k{font-size:12px;color:var(--mut);font-weight:500}
  .kpi .v{font-size:27px;font-weight:800;color:var(--ink);margin-top:6px;letter-spacing:-.02em}
  .kpi .v small{font-size:13px;font-weight:600;color:var(--mut);margin-left:3px}
  .kpi .s{font-size:11.5px;color:var(--good);margin-top:3px}

  .grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
  @media(max-width:980px){.grid{grid-template-columns:1fr}.kpis{grid-template-columns:repeat(2,1fr)}}
  .card{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:14px 16px 8px}
  .card.wide{grid-column:1/-1}
  .card h3{font-size:14px;font-weight:600;color:var(--ink);margin-bottom:2px;padding-left:8px;border-left:3px solid var(--data)}
  .card .sub{font-size:11px;color:var(--faint);margin-bottom:8px;padding-left:8px}
  .chart{width:100%;height:330px}
  .chart.tall{height:380px}
  footer{margin-top:18px;color:var(--faint);font-size:11px;text-align:center;line-height:1.6}

  .m-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
  .m-card{background:var(--bg);border:1px solid var(--line);border-radius:14px;padding:10px 12px 6px}
  .m-card h4{font-size:12.5px;color:var(--ink);font-weight:600;margin-bottom:4px;padding-left:8px;border-left:3px solid var(--data2)}
  .m-chart{width:100%;height:270px}
  .m-table{width:100%;border-collapse:collapse;font-size:12px}
  .m-table th,.m-table td{border:1px solid var(--line);padding:7px 10px;text-align:left}
  .m-table th{background:#22304a;color:#eef2ff;font-weight:600}
  .m-table .r{text-align:right}
  .m-table tr.tot td{background:#3a3118;font-weight:700;color:#fde68a}
  @media(max-width:760px){.m-grid{grid-template-columns:1fr}}

  /* 标签页 */
  .tabs{display:flex;gap:6px;margin:14px 0 4px;border-bottom:1px solid var(--line);padding-bottom:10px}
  .tab{padding:8px 22px;border-radius:10px;cursor:pointer;font-size:13.5px;font-weight:600;
    color:var(--mut);background:var(--card);border:1px solid var(--line);transition:.18s;user-select:none}
  .tab:hover{color:var(--data)}
  .tab.active{background:var(--hero);color:#fff;border-color:var(--hero)}
  .tab.active.fc{background:var(--accent);border-color:var(--accent);color:#1e293b}
  .view{display:none}
  .view.active{display:block}

  /* 预测页 */
  .fc-kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:16px 0}
  @media(max-width:980px){.fc-kpis{grid-template-columns:repeat(2,1fr)}}
  .fc-kpi{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:16px 18px}
  .fc-kpi .k{font-size:12px;color:var(--mut);font-weight:500}
  .fc-kpi .v{font-size:25px;font-weight:800;color:var(--ink);margin-top:6px;letter-spacing:-.02em}
  .fc-kpi .v small{font-size:13px;font-weight:600;color:var(--mut);margin-left:3px}
  .fc-kpi .s{font-size:11.5px;color:var(--mut);margin-top:3px}
  .sms-card{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:18px 20px;margin-top:14px}
  .sms-card h3{font-size:14px;font-weight:600;color:var(--ink);margin-bottom:10px;padding-left:8px;border-left:3px solid var(--accent)}
  .sms-card .sms{font-size:12.8px;color:var(--txt);line-height:1.9;white-space:pre-wrap}
  .sms-card .sms b{color:var(--accent);font-weight:700}
  .note{font-size:11px;color:var(--faint);margin-top:8px;line-height:1.7}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>陕西高速延长石油有限责任公司非油销售汇总驾驶舱 · 图表视图</h1>
    <p>联动仪表盘 · 点击站点可查看该站品类明细（数据取自驾驶舱按月汇总/图表数据源）</p>
    <p class="byline">Design by 刘龙伟 · Email: liulongwei@126.com</p>
    <div class="upd" id="upd"></div>
  </header>

  <div class="tabs" id="tabs">
    <div class="tab active" data-view="overview">总览</div>
    <div class="tab fc" data-view="forecast">预测</div>
  </div>

  <!-- ═══ 总览视图 ═══ -->
  <div class="view active" id="view-overview">
  <div class="stations" id="stations"></div>

  <div class="kpis" id="kpis"></div>

  <div class="grid">
    <div class="card"><h3 id="t1">各站任务完成排行 · 年任务 vs 累计完成（含烟草）</h3>
      <div class="sub">万元 · 深色=当前选中站 · 点击柱体可切换选中并查看明细</div><div id="c1" class="chart"></div></div>
    <div class="card"><h3 id="t2">各站任务完成率（含烟草）</h3>
      <div class="sub">累计完成 ÷ 年任务 · 深色=当前选中站</div><div id="c2" class="chart"></div></div>
    <div class="card"><h3 id="t9">毛利贡献额排名（含烟草 · 年度累计）</h3>
      <div class="sub">万元 · 各站累计毛利降序</div><div id="c9" class="chart"></div></div>
    <div class="card"><h3 id="t10">毛利率排名（含烟草 · 年度累计）</h3>
      <div class="sub">% · 各站累计毛利率降序 · 红线=23.5% 公司红线</div><div id="c10" class="chart"></div></div>
    <div class="card wide"><h3 id="t3">各站分会计月销售趋势（含烟草金额）</h3>
      <div class="sub">元 · 选中单站时该站加粗高亮、其余淡显；选「全公司」则均衡展示</div><div id="c3" class="chart tall"></div></div>
    <div class="card wide"><h3 id="t7">近30天日变化 · 非油销售额（含烟草）</h3>
      <div class="sub">元/日 · 最新30天 · 选「全公司」看公司合计，选中站看该站</div><div id="c7" class="chart"></div></div>
    <div class="card wide"><h3 id="t8">近30天日变化 · 非油毛利（含烟草）</h3>
      <div class="sub">元/日 · 最新30天 · 选「全公司」看公司合计，选中站看该站</div><div id="c8" class="chart"></div></div>
    <div class="card wide"><h3 id="t4">全公司月度非油金额与环比（含烟草）</h3>
      <div class="sub">柱=月度金额(元) · 线=环比(%)</div><div id="c4" class="chart"></div></div>
    <div class="card"><h3 id="t5">金额构成（含烟草 · 年度累计，万元）</h3>
      <div class="sub">汽车用品 / 便利百货 / 香烟零售 / 烟草批发 / 咖啡</div><div id="c5" class="chart"></div></div>
    <div class="card"><h3 id="t6">双口径对比（金额 / 毛利，万元）</h3>
      <div class="sub">含烟草 vs 不含烟草</div><div id="c6" class="chart"></div></div>
  </div>

  <div class="card wide" style="margin-top:14px">
    <h3 id="detailTitle">品类构成与明细</h3>
    <div class="sub" id="detailSub" style="margin-bottom:10px">5 品类 × 12 会计月 金额/毛利堆叠、合计表（品类构成见上方「金额构成」卡）</div>
    <div class="m-grid">
      <div class="m-card"><h4 id="dAmtH">品类 × 12 月 金额（元）</h4><div id="dAmt" class="m-chart" style="height:280px"></div></div>
      <div class="m-card"><h4 id="dProfH">品类 × 12 月 毛利（元）</h4><div id="dProf" class="m-chart" style="height:280px"></div></div>
      <div class="m-card" style="grid-column:1/-1"><h4>金额 / 毛利 年度合计（万元）</h4>
        <div style="overflow:auto"><table class="m-table" id="dTable"></table></div></div>
    </div>
  </div>
  </div>

  <!-- ═══ 预测视图 ═══ -->
  <div class="view" id="view-forecast">
    <div class="fc-kpis" id="fcKpis"></div>
    <div class="card wide">
      <h3>全年预测 · 双情景对比（毛利红线 23.5% vs 24%）</h3>
      <div class="sub">单位：万元 · 23.5%=守住公司整体毛利红线情景 · 24%=提高红线情景（烟草去化更保守）</div>
      <div id="fc1" class="chart tall" style="height:360px"></div>
    </div>
    <div class="card wide" style="margin-top:14px">
      <h3>销量任务超额（双情景）</h3>
      <div class="sub">全年任务 1130 万元 · 超额率 %</div>
      <div id="fc2" class="chart" style="height:220px"></div>
    </div>
    <div class="sms-card">
      <h3>预测短信原文（取自驾驶舱「目标达成」）</h3>
      <div class="sms" id="fcSms"></div>
    </div>
  </div>

</div>

<script>
const DATA = __DATA__;
const CAT_DETAIL = DATA.composition.cats;   // 5 品类（不含非油合计）
const C = { ink:'#eef2ff', txt:'#d7e0f0', mut:'#8fa1c5', faint:'#64748b', grid:'#2b3752',
            data:'#3b82f6', data2:'#60a5fa', hero:'#ffffff', sheet:'#162032' };
const PALETTE = ['#ef4444','#3b82f6','#facc15','#06b6d4','#ec4899','#22c55e','#a855f7','#f97316',
                 '#14b8a6','#eab308','#8b5cf6','#10b981','#f43f5e','#0ea5e9','#d946ef'];
const MODAL_PALETTE = ['#ef4444','#3b82f6','#facc15','#06b6d4','#ec4899','#a855f7'];
const fmtW = v => (v==null?'—':(v/10000).toFixed(2)+'万');
const sum12 = arr => (arr||[]).reduce((a,b)=>a+(b||0),0);
const charts = [];
function mk(id, opt){ const el=document.getElementById(id);
  const c=echarts.init(el,null,{renderer:'canvas'}); c.setOption(opt); charts.push(c); return c; }
const num2 = v => (typeof v === 'number') ? Number(v.toFixed(2)) : v;   // 图表数字统一保留两位小数
const fmt2 = v => (typeof v === 'number') ? v.toLocaleString('zh-CN',{minimumFractionDigits:2, maximumFractionDigits:2}) : v; // 千分位+两位小数
const baseTip = {trigger:'axis', backgroundColor:'#ffffff', borderWidth:0, padding:[9,13],
  textStyle:{color:'#111827', fontFamily:'Inter', fontSize:12},
  valueFormatter: v => Array.isArray(v) ? v.map(num2) : num2(v)};
const tipItem = {trigger:'item', backgroundColor:'#ffffff', borderWidth:0, padding:[9,13],
  textStyle:{color:'#111827', fontFamily:'Inter', fontSize:12},
  valueFormatter: v => num2(v)};
const baseGrid = {left:54,right:22,top:30,bottom:46};
const cat = d => ({type:'category',data:d,axisLine:{lineStyle:{color:C.grid}},
  axisTick:{show:false}, axisLabel:{color:C.mut,fontSize:11}});
const val = (name,fmt) => ({type:'value',name:name,nameTextStyle:{color:C.mut,fontSize:10},
  axisLine:{show:false},splitLine:{lineStyle:{color:C.grid}},axisLabel:{color:C.mut,fontSize:11,formatter:fmt}});

let SEL = '全公司';

// ---------- KPI ----------
function renderKPI(){
  const t = DATA.totals, task = DATA.taskTotal;
  const stObj = SEL === '全公司' ? null : DATA.stations.find(s => s.name === SEL);
  const comp   = stObj ? stObj.compTob   : t.tobAmount;
  const rate   = stObj ? stObj.rateTob*100 : (t.tobAmount/task*100);
  const profit = stObj ? Math.round(sum12(DATA.monthly.tobProfit[SEL])/1e2)/100 : t.tobProfit;
  const profRate = stObj ? (comp ? profit/comp*100 : 0) : (t.tobProfit/t.tobAmount*100);
  const yearTask  = stObj ? stObj.yearTask : task;
  const subt1 = stObj ? `年任务 ${yearTask}万 · 进度 ${rate.toFixed(2)}%`
                       : `全年任务 ${task}万 · 进度 ${(t.tobAmount/task*100).toFixed(2)}%`;
  const cards = [
    {k: stObj ? '本站累计完成(含烟草)' : '累计完成(含烟草)',
      v: comp==null?'—':comp.toFixed(2), u:'万', s: subt1},
    {k: '完成率(含烟草)', v: rate==null?'—':rate.toFixed(2), u:'%',
      s: rate>=100 ? '已达标 ✔' : ((100-rate).toFixed(2)+'% 待完成')},
    {k: stObj ? '本站累计毛利(含烟草)' : '累计毛利(含烟草)',
      v: profit==null?'—':profit.toFixed(2), u:'万',
      s: '毛利率 ' + profRate.toFixed(2) + '%'},
    {k: stObj ? '本站年任务' : '全年任务',
      v: yearTask, u:'万', s: stObj ? '本站任务盘（不含烟草去化）' : '烟草去化已计入'},
  ];
  document.getElementById('kpis').innerHTML = cards.map(c=>
    `<div class="kpi"><div class="k">${c.k}</div><div class="v">${c.v}<small>${c.u}</small></div><div class="s">${c.s}</div></div>`).join('');
}

// ---------- 图1 任务（公司=排行；选中站=该站任务进度） ----------
function renderRank(){
  const t = document.getElementById('t1');
  if(SEL === '全公司'){
    t.textContent = '各站任务完成排行 · 年任务 vs 累计完成（含烟草）';
    const st = DATA.stations.slice().sort((a,b)=>b.compTob-a.compTob);
    const names = st.map(s=>s.name), task = st.map(s=>s.yearTask), done = st.map(s=>s.compTob);
    const colors = st.map(s=> s.name===SEL ? C.hero : C.data2);
    mk('c1',{tooltip:baseTip,legend:{data:['年任务','累计完成'],textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600},top:2},
      grid:baseGrid, xAxis:cat(names), yAxis:val('万元'),
      series:[
        {name:'年任务',type:'bar',data:task,barWidth:13,itemStyle:{color:'#475569',borderRadius:[4,4,0,0]}},
        {name:'累计完成',type:'bar',data:done,barWidth:13,
          itemStyle:{color:p=>colors[p.dataIndex]!==undefined?colors[p.dataIndex]:C.data,borderRadius:[4,4,0,0]},
          label:{show:true,position:'top',color:C.mut,fontSize:9,formatter:p=>Number(p.value).toFixed(2)}}
      ]});
    return;
  }
  const s = DATA.stations.find(x=>x.name===SEL);
  t.textContent = SEL + ' · 任务进度（年任务 vs 累计完成，万元）';
  mk('c1',{tooltip:baseTip, grid:baseGrid, xAxis:cat(['年任务','累计完成']), yAxis:val('万元'),
    series:[{type:'bar',data:[s.yearTask, s.compTob],barWidth:44,
      itemStyle:{color:p=>p.dataIndex===0?'#475569':C.hero,borderRadius:[4,4,0,0]},
      label:{show:true,position:'top',color:C.mut,fontSize:10,formatter:p=>Number(p.value).toFixed(2)}}]});
}
// ---------- 图2 完成率（公司=排行；选中站=含/不含烟草对比） ----------
function renderRate(){
  const t = document.getElementById('t2');
  if(SEL === '全公司'){
    t.textContent = '各站任务完成率（含烟草）';
    const st = DATA.stations.slice().sort((a,b)=>b.rateTob-a.rateTob);
    const names = st.map(s=>s.name);
    const rate = st.map(s=>+(s.rateTob*100).toFixed(2));
    const colors = st.map(s=> s.name===SEL ? C.hero : C.data2);
    mk('c2',{tooltip:{...baseTip,formatter:p=>p[0].name+'：'+Number(p[0].value).toFixed(2)+'%'},grid:baseGrid,
      xAxis:cat(names), yAxis:val('完成率',v=>v+'%'),
      series:[{type:'bar',data:rate.map((v,i)=>({value:v,itemStyle:{color:colors[i],borderRadius:[4,4,0,0]}})),
        barWidth:18, label:{show:true,position:'top',color:C.mut,fontSize:9,formatter:p=>Number(p.value).toFixed(2)+'%'}}]});
    return;
  }
  const s = DATA.stations.find(x=>x.name===SEL);
  t.textContent = SEL + ' · 完成率（含烟草 vs 不含烟草）';
  mk('c2',{tooltip:{...baseTip,formatter:p=>p[0].name+'：'+Number(p[0].value).toFixed(2)+'%'},grid:baseGrid,
    xAxis:cat(['含烟草','不含烟草']), yAxis:val('完成率',v=>v+'%'),
    series:[{type:'bar',data:[+(s.rateTob*100).toFixed(2), +(s.rateNon*100).toFixed(2)],barWidth:44,
      itemStyle:{color:p=>p.dataIndex===0?C.hero:C.data2,borderRadius:[4,4,0,0]},
      label:{show:true,position:'top',color:C.mut,fontSize:10,formatter:p=>Number(p.value).toFixed(2)+'%'}}]});
}
// ---------- 图3 趋势(联动) ----------
function renderTrend(){
  const t = document.getElementById('t3');
  t.textContent = SEL==='全公司' ? '各站分会计月销售趋势（含烟草金额）' : SEL + ' · 月度销售趋势（含烟草金额）';
  const months = DATA.company.monthLabels;
  const m = DATA.monthly.tob;
  const series = DATA.stations.map((s,i)=>{
    const isSel = (SEL===s.name);
    const dim = (SEL!=='全公司' && !isSel);
    return {name:s.name, type:'line', smooth:true, symbol:'none',
      data:m[s.name],
      lineStyle:{width:isSel?3.4:1.4, color:isSel?C.hero:PALETTE[i%PALETTE.length], opacity:dim?0.18:1},
      itemStyle:{color:isSel?C.hero:PALETTE[i%PALETTE.length], opacity:dim?0.18:1},
      z:isSel?10:1, emphasis:{focus:'series'}};
  });
  mk('c3',{tooltip:baseTip, legend:{show:false}, grid:{left:60,right:24,top:18,bottom:40},
    xAxis:cat(months), yAxis:val('元'), series});
}
// ---------- 图4 月度（公司=公司月度；选中站=该站月度） ----------
function renderCompany(){
  const t = document.getElementById('t4');
  const months = DATA.company.monthLabels;
  if(SEL === '全公司'){
    t.textContent = '全公司月度非油金额与环比（含烟草）';
    const c = DATA.company;
    mk('c4',{tooltip:baseTip, legend:{data:['含烟草金额','含烟草环比'],textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600},top:2},
      grid:Object.assign({},baseGrid,{right:56}), xAxis:cat(months),
      yAxis:[val('元'),{type:'value',name:'环比',nameTextStyle:{color:C.mut,fontSize:10},
        axisLine:{show:false},splitLine:{show:false},axisLabel:{color:C.mut,fontSize:11,formatter:v=>(v*100).toFixed(0)+'%'}}],
      series:[
        {name:'含烟草金额',type:'bar',data:c.tob,barWidth:20,itemStyle:{color:C.data,borderRadius:[4,4,0,0]}},
        {name:'含烟草环比',type:'line',yAxisIndex:1,smooth:true,data:c.tobMom,
          lineStyle:{width:2,color:C.accent},itemStyle:{color:C.accent}}
      ]});
    return;
  }
  t.textContent = SEL + ' · 月度非油金额与环比（含烟草）';
  const arr = DATA.monthly.tob[SEL] || [];
  const mom = arr.map((v,i)=>{ if(i===0) return null; const p=arr[i-1]; if(!p) return null; return (v-p)/p; });
  mk('c4',{tooltip:baseTip, legend:{data:['月度金额','环比'],textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600},top:2},
    grid:Object.assign({},baseGrid,{right:56}), xAxis:cat(months),
    yAxis:[val('元'),{type:'value',name:'环比',nameTextStyle:{color:C.mut,fontSize:10},
      axisLine:{show:false},splitLine:{show:false},axisLabel:{color:C.mut,fontSize:11,formatter:v=>v==null?'':(v*100).toFixed(0)+'%'}}],
    series:[
      {name:'月度金额',type:'bar',data:arr,barWidth:20,itemStyle:{color:C.data,borderRadius:[4,4,0,0]}},
      {name:'环比',type:'line',yAxisIndex:1,smooth:true,data:mom,
        lineStyle:{width:2,color:C.accent},itemStyle:{color:C.accent}}
    ]});
}
// ---------- 图5 构成（公司=全公司5类；选中站=该站5类） ----------
function renderComp(){
  const t = document.getElementById('t5');
  let names, vals, profits;
  if(SEL === '全公司'){
    t.textContent = '金额构成（含烟草 · 年度累计，万元）';
    names = DATA.composition.cats; vals = DATA.composition.tob;
    // 公司级各品类毛利 = 各站该品类年度毛利合计
    profits = names.map(c=> Math.round(
      DATA.stations.reduce((s,st)=> s+sum12(DATA.categoryDetail.profit[c][st.name]||[]), 0)/1e2)/100);
  }else{
    t.textContent = SEL + ' · 品类构成（年度累计，万元）';
    names = CAT_DETAIL;
    vals = CAT_DETAIL.map(c=> Math.round(sum12(DATA.categoryDetail.amount[c][SEL])/1e2)/100);
    profits = CAT_DETAIL.map(c=> Math.round(sum12(DATA.categoryDetail.profit[c][SEL])/1e2)/100);
  }
  const data = names.map((n,i)=>{
    const amt = vals[i], prof = profits[i];
    const rate = amt ? (prof/amt*100) : 0;
    return {name:n, value:amt, profit:prof, rate:rate};
  });
  mk('c5',{tooltip:{...tipItem, formatter:p=>{
      const d=p.data;
      return d.name+'<br>金额：'+fmt2(d.value)+' 万<br>毛利：'+fmt2(d.profit)+' 万<br>毛利率：'+Number(d.rate).toFixed(2)+'%';
    }},
    legend:{orient:'vertical',left:'left',textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600}},
    color:MODAL_PALETTE,
    series:[{type:'pie',radius:['42%','68%'],center:['62%','54%'],
      label:{color:C.txt,fontSize:10.5,formatter:p=>{
        const d=p.data;
        return d.name+'\n'+num2(d.value)+'万 · 毛利 '+num2(d.profit)+'万 ('+Number(d.rate).toFixed(1)+'%)';
      }},
      labelLine:{lineStyle:{color:C.faint}},
      data:data}]});
}
// ---------- 图6 双口径（公司=公司；选中站=该站） ----------
function renderDual(){
  const t = document.getElementById('t6');
  let d;
  if(SEL === '全公司'){
    t.textContent = '双口径对比（金额 / 毛利，万元）';
    d = DATA.dual;
  }else{
    t.textContent = SEL + ' · 双口径对比（金额 / 毛利，万元）';
    const m = DATA.monthly;
    d = {cats:['金额','毛利'],
      non:[ Math.round(sum12(m.non[SEL])/1e2)/100, Math.round(sum12(m.nonProfit[SEL])/1e2)/100 ],
      tob:[ Math.round(sum12(m.tob[SEL])/1e2)/100, Math.round(sum12(m.tobProfit[SEL])/1e2)/100 ]};
  }
  mk('c6',{tooltip:baseTip, legend:{data:['不含烟草','含烟草'],textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600},top:2},
    grid:baseGrid, xAxis:cat(d.cats), yAxis:val('万元'),
    series:[
      {name:'不含烟草',type:'bar',data:d.non,barWidth:34,itemStyle:{color:C.data2,borderRadius:[4,4,0,0]}},
      {name:'含烟草',type:'bar',data:d.tob,barWidth:34,itemStyle:{color:C.data,borderRadius:[4,4,0,0]}}
    ]});
}

// ---------- 近30天日变化(联动) ----------
function renderDaily30(){
  const d = DATA.daily30 || {dates:[], amount:{}, profit:{}};
  if(!d.dates.length){ return; }
  const key = SEL === '全公司' ? '合计' : SEL;
  const amtArr = d.amount[key] || [];
  const profArr = d.profit[key] || [];
  const t7 = document.getElementById('t7');
  const t8 = document.getElementById('t8');
  t7.textContent = '近30天日变化 · 非油销售额（含烟草）' + (SEL==='全公司' ? '' : ' · ' + SEL);
  t8.textContent = '近30天日变化 · 非油毛利（含烟草）' + (SEL==='全公司' ? '' : ' · ' + SEL);
  // 7日均线
  const ma7 = arr => arr.map((_,i)=>{
    if(i<6) return null;
    let s=0; for(let j=i-6;j<=i;j++) s+=(arr[j]||0);
    return Math.round(s/7*100)/100;
  });
  const mkDaily = (id, arr, color) => {
    mk(id,{tooltip:{...baseTip,formatter:ps=>{
        const p=ps[0]; let html=p.name+'<br>';
        ps.forEach(x=>{ html+=x.marker+x.seriesName+'：'+fmt2(x.value==null?0:x.value)+' 元<br>'; });
        return html;
      }},
      legend:{data:['日值','7日均线'],textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600},top:2},
      grid:{left:70,right:26,top:34,bottom:46}, xAxis:cat(d.dates),
      yAxis:val('元'),
      series:[
        {name:'日值',type:'line',smooth:true,symbol:'circle',symbolSize:4,data:arr,
          lineStyle:{width:2,color:color},itemStyle:{color:color},areaStyle:{color:color,opacity:0.08}},
        {name:'7日均线',type:'line',smooth:true,symbol:'none',data:ma7(arr),
          lineStyle:{width:1.6,color:'#fbbf24',type:'dashed'},itemStyle:{color:'#fbbf24'}}
      ]});
  };
  mkDaily('c7', amtArr, C.data);
  mkDaily('c8', profArr, C.data2);
}

// ---------- 毛利贡献额 / 毛利率 排名榜(公司视角,横向条形) ----------
function renderRank2(){
  const prof = DATA.stations.map(s=>({
    name:s.name,
    amt: Math.round(sum12(DATA.monthly.tob[s.name])/1e2)/100,          // 万元
    profit: Math.round(sum12(DATA.monthly.tobProfit[s.name])/1e2)/100, // 万元
  }));
  prof.forEach(p=>{ p.rate = p.amt ? +(p.profit/p.amt*100).toFixed(2) : 0; });
  const byProfit = prof.slice().sort((a,b)=>b.profit-a.profit);
  const byRate   = prof.slice().sort((a,b)=>b.rate-a.rate);

  // 毛利贡献额榜(横向)
  mk('c9',{tooltip:{...tipItem,valueFormatter:v=>num2(v)+' 万'},
    grid:{left:70,right:44,top:16,bottom:26},
    xAxis:{type:'value',axisLine:{show:false},splitLine:{lineStyle:{color:C.grid}},
      axisLabel:{color:C.mut,fontSize:11,formatter:v=>v+'万'}},
    yAxis:{type:'category',inverse:true,data:byProfit.map(x=>x.name),
      axisLine:{show:false},axisTick:{show:false},axisLabel:{color:C.txt,fontSize:11.5}},
    series:[{type:'bar',data:byProfit.map((x,i)=>({
        value:x.profit, itemStyle:{color:i===0?'#fbbf24':(i===1?'#60a5fa':'#3b82f6'),
          borderRadius:[0,4,4,0]}})),
      barWidth:16,
      label:{show:true,position:'right',color:C.txt,fontSize:10.5,formatter:p=>num2(p.value)+'万'}}]});

  // 毛利率榜(横向 + 23.5%红线)
  mk('c10',{tooltip:{...tipItem,valueFormatter:v=>num2(v)+'%'},
    grid:{left:70,right:44,top:16,bottom:26},
    xAxis:{type:'value',axisLine:{show:false},splitLine:{lineStyle:{color:C.grid}},
      axisLabel:{color:C.mut,fontSize:11,formatter:v=>v+'%'},
      max: Math.max(...byRate.map(x=>x.rate), 23.5)*1.12},
    yAxis:{type:'category',inverse:true,data:byRate.map(x=>x.name),
      axisLine:{show:false},axisTick:{show:false},axisLabel:{color:C.txt,fontSize:11.5}},
    series:[{type:'bar',data:byRate.map((x,i)=>({
        value:x.rate,
        itemStyle:{color: x.rate>=23.5 ? '#34d399' : '#ef4444', borderRadius:[0,4,4,0]}})),
      barWidth:16,
      label:{show:true,position:'right',color:C.txt,fontSize:10.5,formatter:p=>num2(p.value)+'%'},
      markLine:{symbol:'none',data:[{xAxis:23.5,lineStyle:{color:'#fbbf24',type:'dashed',width:1.5},
        label:{formatter:'红线 23.5%',color:'#fbbf24',fontSize:10,position:'end'}}]}}]});
}

// ---------- 品类明细（内嵌到主页，4 块：金额堆叠/毛利堆叠/构成饼/合计表） ----------
function renderDetail(){
  const isCo = SEL === '全公司';
  const t = document.getElementById('detailTitle');
  const sub = document.getElementById('detailSub');
  const aH = document.getElementById('dAmtH');
  const pH = document.getElementById('dProfH');
  const months = DATA.company.monthLabels;
  const amt = {}, prof = {};
  if (isCo) {
    t.textContent = '全公司 · 品类构成与明细';
    sub.textContent = '5 品类 × 12 会计月 金额/毛利堆叠（公司合计）、年度构成、合计表';
    aH.textContent = '品类 × 12 月 金额（元，公司合计）';
    pH.textContent = '品类 × 12 月 毛利（元，公司合计）';
    for (const c of CAT_DETAIL) {
      const aArr = new Array(12).fill(0), pArr = new Array(12).fill(0);
      for (const st of DATA.stations) {
        const a = DATA.categoryDetail.amount[c][st.name] || [];
        const p = DATA.categoryDetail.profit[c][st.name] || [];
        for (let m = 0; m < 12; m++) { aArr[m] += (a[m] || 0); pArr[m] += (p[m] || 0); }
      }
      amt[c] = aArr; prof[c] = pArr;
    }
  } else {
    t.textContent = SEL + ' · 品类构成与明细';
    sub.textContent = '5 品类 × 12 会计月 金额/毛利堆叠、年度构成、合计表';
    aH.textContent = '品类 × 12 月 金额（元）';
    pH.textContent = '品类 × 12 月 毛利（元）';
    for (const c of CAT_DETAIL) {
      amt[c]  = DATA.categoryDetail.amount[c][SEL] || [];
      prof[c] = DATA.categoryDetail.profit[c][SEL]  || [];
    }
  }
  mk('dAmt',{tooltip:baseTip, legend:{textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600}, top:6},
    grid:{left:54,right:22,top:38,bottom:46},
    xAxis:cat(months), yAxis:val('元'),
    series:CAT_DETAIL.map((c,i)=>({name:c,type:'bar',stack:'a',data:amt[c],barMaxWidth:24,
      itemStyle:{color:MODAL_PALETTE[i]}}))});
  mk('dProf',{tooltip:baseTip, legend:{textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600}, top:6},
    grid:{left:54,right:22,top:38,bottom:46},
    xAxis:cat(months), yAxis:val('元'),
    series:CAT_DETAIL.map((c,i)=>({name:c,type:'bar',stack:'p',data:prof[c],barMaxWidth:24,
      itemStyle:{color:MODAL_PALETTE[i]}}))});
  const rows = CAT_DETAIL.map(c=>{
    const a = Math.round(sum12(amt[c])/1e2)/100, p = Math.round(sum12(prof[c])/1e2)/100;
    return `<tr><td>${c}</td><td class="r">${fmt2(a)}</td><td class="r">${fmt2(p)}</td></tr>`;
  }).join('');
  const ta = Math.round(CAT_DETAIL.reduce((s,c)=>s+sum12(amt[c]),0)/1e2)/100;
  const tp = Math.round(CAT_DETAIL.reduce((s,c)=>s+sum12(prof[c]),0)/1e2)/100;
  document.getElementById('dTable').innerHTML =
    `<tr><th>品类</th><th>金额（万元）</th><th>毛利（万元）</th></tr>` + rows +
    `<tr class="tot"><td>合计</td><td class="r">${fmt2(ta)}</td><td class="r">${fmt2(tp)}</td></tr>`;
}

// ---------- 渲染总入口 ----------
function renderAll(){
  charts.forEach(c=>c.dispose()); charts.length=0;
  renderKPI(); renderRank(); renderRate(); renderRank2(); renderTrend(); renderDaily30(); renderCompany(); renderComp(); renderDual();
  renderDetail();
  document.getElementById('upd').textContent = '数据生成于 ' + DATA.generatedAt + ' · 当前视角：' + SEL;
}

// ---------- 站点选择器 ----------
function buildChips(){
  const box = document.getElementById('stations');
  const all = document.createElement('div');
  all.className='chip all'+(SEL==='全公司'?' active':'');
  all.textContent='全公司'; all.onclick=()=>{SEL='全公司';syncChips();renderAll();};
  box.appendChild(all);
  DATA.stations.forEach(s=>{
    const c=document.createElement('div');
    c.className='chip'+(SEL===s.name?' active':''); c.textContent=s.name;
    c.onclick=()=>{SEL=s.name;syncChips();renderAll();};
    box.appendChild(c);
  });
}
function syncChips(){ document.querySelectorAll('#stations .chip').forEach(c=>{
  c.classList.toggle('active', (c.classList.contains('all')&&SEL==='全公司') || c.textContent===SEL); }); }

// ---------- 标签页切换 ----------
function switchView(name){
  document.querySelectorAll('#tabs .tab').forEach(t=>t.classList.toggle('active', t.dataset.view===name));
  document.querySelectorAll('.view').forEach(v=>v.classList.toggle('active', v.id==='view-'+name));
  if(name==='forecast'){ renderForecast(); }
  setTimeout(()=>charts.forEach(c=>c.resize()), 60);
}

// ---------- 预测视图 ----------
function renderForecast(){
  const fc = DATA.forecast || {};
  const k = fc.kpis || {};
  const ex = fc.extra || {};
  // KPI
  const cards = [
    {k:'累计含去化销售额', v:k['累计含去化销售额万']==null?'—':Number(k['累计含去化销售额万']).toFixed(2), u:'万', s:'含烟草批发去化口径'},
    {k:'自然口径累计', v:k['自然口径累计万']==null?'—':Number(k['自然口径累计万']).toFixed(2), u:'万', s:'不含烟草批发的自然经营口径'},
    {k:'整体毛利率', v:k['整体毛利率']==null?'—':Number(k['整体毛利率']).toFixed(2), u:'%', s:'当前累计毛利 ÷ 累计销售额'},
    {k:'华山站停业扣除', v:k['华山扣除万']==null?'—':Number(k['华山扣除万']).toFixed(2), u:'万', s:'10/15–12/25 停业预测损失'},
  ];
  document.getElementById('fcKpis').innerHTML = cards.map(c=>
    `<div class="fc-kpi"><div class="k">${c.k}</div><div class="v">${c.v}<small>${c.u}</small></div><div class="s">${c.s}</div></div>`).join('');

  // 双情景对比图
  const sc = fc.scenarios || {labels:[], A:[], B:[]};
  mk('fc1',{tooltip:{...baseTip,valueFormatter: v=>num2(v)+' 万'},
    legend:{data:['23.5%红线','24%红线'],textStyle:{color:'#dbe3f0',fontSize:12,fontWeight:600},top:2},
    grid:{left:56,right:24,top:36,bottom:40}, xAxis:cat(sc.labels), yAxis:val('万元'),
    series:[
      {name:'23.5%红线',type:'bar',data:sc.A,barWidth:26,itemStyle:{color:'#34d399',borderRadius:[4,4,0,0]},
        label:{show:true,position:'top',color:C.mut,fontSize:9.5,formatter:p=>Number(p.value).toFixed(1)}},
      {name:'24%红线',type:'bar',data:sc.B,barWidth:26,itemStyle:{color:'#fbbf24',borderRadius:[4,4,0,0]},
        label:{show:true,position:'top',color:C.mut,fontSize:9.5,formatter:p=>Number(p.value).toFixed(1)}}
    ]});

  // 销量超额
  const ov = fc.overrun || {};
  mk('fc2',{tooltip:{...baseTip,formatter:p=>p[0].name+'：'+Number(p[0].value).toFixed(2)+'%'},
    grid:{left:56,right:24,top:24,bottom:34}, xAxis:cat(['23.5%红线','24%红线']),
    yAxis:val('超额率',v=>v+'%'),
    series:[{type:'bar',data:[ov.A, ov.B],barWidth:52,
      itemStyle:{color:p=>p.dataIndex===0?'#34d399':'#fbbf24',borderRadius:[4,4,0,0]},
      label:{show:true,position:'top',color:C.mut,fontSize:11,formatter:p=>Number(p.value).toFixed(2)+'%'}}]});

  // 短信原文
  const sms = fc.sms || '';
  const lines = sms.split('\n').filter(Boolean);
  const esc = s => s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  document.getElementById('fcSms').innerHTML = lines.map((ln,i)=>`<div>${esc(ln)}</div>`).join('')
    + `<div class="note">数据截至 ${fc.asOf||'—'} · 预测截止 ${fc.predEnd||'—'} · 剩余 ${fc.remainingMonths??'—'} 个月 · 当前香烟库存 ${ex['当前香烟库存万']??'—'}万（库存约束去化上限 ${ex['库存约束去化上限万']??'—'}万${ex['库存是否封顶']===1?'，已封顶':''}）</div>`;
}

// 图1/图2 柱体点击 -> 选中该站 (联动)
function bindChartClick(){
  ['c1','c2'].forEach(id=>{
    const c = echarts.getInstanceByDom(document.getElementById(id));
    if(c) c.on('click', p=>{ if(p.name && DATA.stations.some(s=>s.name===p.name)){ SEL=p.name; syncChips(); renderAll(); }});
  });
}

buildChips();
renderAll();
bindChartClick();
document.querySelectorAll('#tabs .tab').forEach(t=>{
  t.addEventListener('click', ()=>switchView(t.dataset.view));
});
window.addEventListener('resize',()=>charts.forEach(c=>c.resize()));
</script>
</body>
</html>
"""


def render(data, out_path=INDEX):
    html = HTML.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


def main():
    if "--from-json" in sys.argv and os.path.exists(DATA_JSON):
        with open(DATA_JSON, encoding="utf-8") as f:
            data = json.load(f)
        print("▶ 仅从 data.json 重渲染 index.html")
    else:
        data = build_data()
        with open(DATA_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
        print("▶ 已从 Excel 提取数据 -> data.json")
    render(data)
    print("✅ 已生成:", INDEX)
    print("   站点", len(data["stations"]), "站 | 趋势", len(data["monthly"]["tob"]),
          "站 | 构成", len(data["composition"]["cats"]), "类 | 品类明细",
          len(data["categoryDetail"]["cats"]), "类×", len(data["categoryDetail"]["amount"]["汽车用品"]), "站×12月")
    print("   公司累计(含烟草):", data["totals"]["tobAmount"], "万 | 毛利:",
          data["totals"]["tobProfit"], "万 | 任务:", data["taskTotal"], "万")


if __name__ == "__main__":
    main()
